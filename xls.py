import openpyxl
from openpyxl.styles import Font


class Excel:
    def __init__(self, registry_path, dir_scan, ws_name, settings):
        font_size = int(settings['font_size'])
        font_name = settings['font_name']
        hyperlink_color = settings['hyperlink_color']
        self.font = Font(name=font_name, size=font_size, underline='single', color=hyperlink_color)
        self.registry_path = registry_path
        self.dir_scan = dir_scan
        self.ws_name = ws_name
        self.wb = openpyxl.open(self.registry_path)
        print(f'Документ {self.registry_path} открыт.')
        if self.ws_name not in self.wb.sheetnames:
            print(f'Лист {self.ws_name} отсутствует в книге.')
        self.ws = self.wb[self.ws_name]
        print(f'Выбран лист {self.ws_name}.')

    def get_column(self):
        list_column = list(self.ws[f'A{i}'].value for i in range(3, self.ws.max_row + 1))
        list_column = list(str(i) for i in list_column if i is not None)
        print('Получен список регистрационных номеров из столбца "А".')
        return list_column

    def create_hyperlinks(self, name, link_name, position):
        self.ws[f'H{position}'] = name
        self.ws[f'H{position}'].hyperlink = f'{self.dir_scan}\\{link_name}'
        self.ws[f'H{position}'].font = self.font

    def save(self):
        self.wb.save(self.registry_path)


if __name__ == '__main__':
    pass